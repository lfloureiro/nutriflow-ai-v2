import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

PORTFIR_VERSION = "7.1-2026"
PORTFIR_WORKBOOK_URL = (
    "https://portfir.insa.min-saude.pt/wp-content/uploads/2025/11/insa_tca.xlsx"
)
PORTFIR_TIMEOUT_SECONDS = 20.0


class PortfirError(ValueError):
    pass


@dataclass(frozen=True)
class PortfirNutrient:
    key: str
    value: Decimal
    unit: str


@dataclass(frozen=True)
class PortfirFoodNutrition:
    code: str
    name: str
    energy_kcal: Decimal
    nutrients: tuple[PortfirNutrient, ...]
    version: str = PORTFIR_VERSION

    @property
    def source_reference(self) -> str:
        return PORTFIR_WORKBOOK_URL


def download_portfir_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    request = Request(
        PORTFIR_WORKBOOK_URL,
        headers={
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "User-Agent": "NutriFlowAI/0.1 nutrition-enrichment",
        },
    )
    try:
        with urlopen(request, timeout=PORTFIR_TIMEOUT_SECONDS) as response:
            data = response.read()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise PortfirError("PortFIR workbook is unavailable.") from exc
    if not data.startswith(b"PK"):
        raise PortfirError("PortFIR did not return an XLSX workbook.")
    path.write_bytes(data)
    return path


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _normalized_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", _text(value).casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _header_key(value: object) -> str | None:
    header = _normalized_header(value)
    if not header:
        return None
    if header in {"id", "codigo", "cod", "codigo alimento", "id alimento"}:
        return "code"
    if header in {"nome", "alimento", "nome alimento", "nome do alimento"}:
        return "name"
    if "energia" in header and "kcal" in header:
        return "energy"
    if "proteina" in header and "g" in header:
        return "protein"
    if ("lipidos" in header or "gordura" in header) and "g" in header:
        if "satur" not in header:
            return "fat"
    if "hidratos" in header and "carbono" in header and "g" in header:
        if "acucar" not in header:
            return "carbohydrate"
    if "fibra" in header and "g" in header:
        return "fiber"
    if "sodio" in header and "mg" in header:
        return "sodium"
    return None


def _header_map(values: tuple[object, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        key = _header_key(value)
        if key is not None and key not in result:
            result[key] = index
    return result


def _find_table(worksheet) -> tuple[int, dict[str, int]] | None:
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=30, values_only=True),
        start=1,
    ):
        mapping = _header_map(tuple(row))
        if "name" in mapping and "energy" in mapping:
            return row_number, mapping
    return None


def _decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = _text(value)
    if not text:
        return None
    if text.startswith("<") or text.casefold() in {"tr", "trace", "n/a", "na", "nd"}:
        return None
    text = text.strip("[]() ").replace(" ", "").replace(",", ".")
    text = re.sub(r"[*a-zA-Z]+$", "", text).strip()
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _value(row: tuple[object, ...], mapping: dict[str, int], key: str) -> object | None:
    index = mapping.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_row(
    row: tuple[object, ...],
    mapping: dict[str, int],
    *,
    row_number: int,
) -> PortfirFoodNutrition | None:
    name = _text(_value(row, mapping, "name"))
    energy = _decimal(_value(row, mapping, "energy"))
    if not name or energy is None or energy < 0:
        return None
    raw_code = _text(_value(row, mapping, "code"))
    code = raw_code or f"row-{row_number}"
    nutrients: list[PortfirNutrient] = []
    for key, unit in (
        ("protein", "g"),
        ("fat", "g"),
        ("carbohydrate", "g"),
        ("fiber", "g"),
        ("sodium", "mg"),
    ):
        amount = _decimal(_value(row, mapping, key))
        if amount is not None and amount >= 0:
            nutrients.append(PortfirNutrient(key=key, value=amount, unit=unit))
    return PortfirFoodNutrition(
        code=code,
        name=name,
        energy_kcal=energy,
        nutrients=tuple(nutrients),
    )


def load_portfir_foods(path: Path) -> tuple[PortfirFoodNutrition, ...]:
    if not path.is_file():
        raise PortfirError(f"PortFIR workbook was not found: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        raise PortfirError("PortFIR workbook could not be opened.") from exc
    try:
        for worksheet in workbook.worksheets:
            table = _find_table(worksheet)
            if table is None:
                continue
            header_row, mapping = table
            foods: list[PortfirFoodNutrition] = []
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                parsed = _parse_row(tuple(row), mapping, row_number=row_number)
                if parsed is not None:
                    foods.append(parsed)
            if foods:
                return tuple(foods)
    finally:
        workbook.close()
    raise PortfirError("PortFIR workbook does not contain a recognizable food table.")
